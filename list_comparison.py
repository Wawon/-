import openpyxl

def compare_names(file1, file2):
    # 加载两个xlsx文件
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # 选择第一个工作表
    sheet1 = wb1.active
    sheet2 = wb2.active

    # 获取人名列表
    names1 = [sheet1.cell(row=i, column=1).value for i in range(1, sheet1.max_row + 1)]
    names2 = [sheet2.cell(row=i, column=1).value for i in range(1, sheet2.max_row + 1)]

    # 比较人名是否存在于另一个文件中
    for name in names1:
        if name in names2:
            print(f"{name} 存在于 {file2}")
        else:
            print(f"{name} 不存在于 {file2}")

# 请将以下文件名替换为您要比较的xlsx文件名
file1 = r"file1的路径和文件名"
file2 = r"file2的路径和文件名"

compare_names(file1, file2)

"""请将file1和file2变量替换为您要比较的xlsx文件名。程序将逐行读取第一个文件中的人名，并检查它们是否存在于第二个文件中。如果存在，程序将输出相应的信息 """
